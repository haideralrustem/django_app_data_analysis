import xlrd
import re
import os
import csv
import pandas as pd
from collections import Counter
import datetime
import numpy as np
import openpyxl

import dateutil.parser  as dparser

import nltk
import pdb
# from scipy import stats
#nltk.download('punkt')
from nltk.tokenize import RegexpTokenizer


import sys
# sys.path.append('C:\\Users\\halrustem\\Documents\\my Python modules')



# ............................................................
def read_excel_file_pd(filename):
  df1 = pd.read_excel(
    filename,
    engine='openpyxl',
  )

  return df1

# .................................

def remove_full_empty_rows(pyxl_sheet):

  new_sht_values = []
  row_idx = 0
  index_first_not_empty = -1
  
  for row in pyxl_sheet:
    none_counts = 0
    for cell in row:
      val = cell.value

      if val == None or val == '':
        # print(val, '\t row_idx', row_idx, '\t len(row)->', len(row))
        none_counts += 1
        
    if none_counts == len(row) and index_first_not_empty== -1:
      print(f'all nones!\t none_counts: {none_counts},\t len(row): {len(row)}, \trow_idx: {row_idx}')
      
    else:
      index_first_not_empty = row_idx
      new_sht_values.append(row)
    row_idx += 1

  return new_sht_values



# .................................
def check_if_date(str_val):
  """range are not checked, just digit postions and pattern of appearance """
  str_val = str(str_val)
  str_val = re.sub(r'\s', '', str_val)
  pat1 = r'\d{1,2}[\-/]\d{1,2}[\-/]\d{2,4}'  # mm-dd-yyyy or dd-mm-yyyy
  pat2 = r'\d{2,4}[\-/]\d{1,2}[\-/]\d{1,2}'  # yyyy-mm-dd or yyyy-mm-dd
  lipat = [pat1, pat2]

  for p in lipat:
    sm = re.search(p, str_val)
    if sm:
      return True

  return False
# ................................

def try_extract_date(text_val):

  date_obj = dparser.parse(text_val, fuzzy=True)
  return date_obj
  # try:
  #   date_obj = dparser.parse(text_val, fuzzy=True)
  #   return date_obj
  # except:
  #   print(f'\tno date could be extraceted! attempted val = {text_val}')
  #   return  text_val

# ...........................

def is_date(string, fuzzy=False):
  """
  Return whether the string can be interpreted as a date.

  :param string: str, string to check for date
  :param fuzzy: bool, ignore unknown tokens in string if True
  """
  try:
    dparse.parse(string, fuzzy=fuzzy)
    return True

  except ValueError:
    return False


# ..................................

def remove_totals_rows(pyxl_sheet):

  words = ['Totals', 'Grand Total' ]
  new_sht_values = []
  row_idx = 0

  for row in pyxl_sheet:
    included_row = True

    for cell in row:
      val = cell.value

      if str(val).strip() in words:
        included_row = False

    # exclude that row
    if included_row:
      new_sht_values.append(row)

  return new_sht_values

# ...............................



# ......................................


def read_excel_file_pyxl(sheet, headers_index=0):

  # print(wb.worksheets)

  result_data = {}

  new_sheet = remove_full_empty_rows(sheet)
  new_sheet = remove_totals_rows(new_sheet)
  
  rows_list = []

  colnames = [t.value for t in new_sheet[headers_index]]
   
  for row in new_sheet[ headers_index+1 : ]:
    converted_row = [cell.value for cell in row]

    rows_list.append(converted_row)


  df = pd.DataFrame(rows_list, columns=colnames)
  return colnames, df, new_sheet, rows_list

# .......................................

def transform_excel_dataframe_to_dictlist(colnames, df, workbook):
  transformed_dictlist = []
  

  return transformed_dict

# ...................................

def read_excel_file(sheet, headers_row_index=0, workbook=None):

  excel_sheet = sheet  # read sheet

  data_dict = {}

  colnames = []  # Reading actual names of excel columns so that we use strings to locate them
  # ASSUMING THAT the first row (index = 0) of Excel sheet is the HEADERS

  # iterate rows
  for row in range(headers_row_index, excel_sheet.nrows):  #0 should be the column headers and not data
    #
    if row > headers_row_index:
      data_dict[row] = {} # initilize empty dict so that data can be captured

    # print('row number : ', row)

    for col_index in range(excel_sheet.ncols):

      cell_val = excel_sheet.cell_value(row, col_index)
      cell_obj = excel_sheet.cell(row, col_index)

      if row == headers_row_index:
        colname = cell_val
        colnames.append(colname)

      elif (row > headers_row_index):
        # print(colnames)
        # print(row, col_index)
        col_key = colnames[col_index]

        if cell_obj.ctype == xlrd.XL_CELL_DATE:
          # print(str(cell_obj).split(':')[0], '\t', cell_obj.ctype)
          cell_val = datetime.datetime(*xlrd.xldate_as_tuple(cell_val, workbook.datemode))

        if cell_val == None:
          cell_val = ''
        data_dict[row][col_key] = cell_val

  return data_dict

# ...........................
# =========================================================

def read_excel_comprehensive(filename, headers_index=0):
  sheet_names = {}

  split_tup = os.path.splitext(filename)
  file_extension = split_tup[-1]

  workbook_dict = {}

  if file_extension == '.xls':

    wb = xlrd.open_workbook(filename)

    s = 0
    for sheet in wb.sheets():
      dat_dict = {}
      dat_dict = read_excel_file(sheet, headers_index, workbook=wb)
      workbook_dict[s] = dat_dict
      sheet_names[s] = sheet.name
      s += 1


  elif file_extension == '.xlsx' :

    wb = openpyxl.load_workbook(filename, read_only=True)

    s=0
    for sheet in wb.worksheets:
      sheet_names[s] = sheet.title

      dat_dict = {}
      colnames, df, sheet, rows_list = read_excel_file_pyxl(sheet, headers_index)


      df_vals_as_list = df.values.tolist()

      r = 1
      for row in df_vals_as_list:
        row_as_dict = {}
        c = 0
        for col in colnames:
          cell_val = row[c]
          if cell_val == None:
            cell_val = ''

          if 'time' in (str(df[col].dtype)) or 'date' in (str(df[col].dtype)):
            cell_val = cell_val.to_pydatetime()

          row_as_dict[col] = cell_val
          c+=1

        dat_dict[r] = row_as_dict
        r += 1

      workbook_dict[s] = dat_dict
      s += 1


  elif file_extension == '.csv':
    s = 0
    for sheet in [0]: # one sheet?
      dat_dict = {}

      df = pd.read_csv(filename)
      colnames = list(df.columns)
      df_vals_as_list = df.values.tolist()
      r = 1
      for row in df_vals_as_list:
        row_as_dict = {}
        c = 0
        for col in colnames:
          cell_val = row[c]
          if cell_val == None:
            cell_val = ''

          if (check_if_date(cell_val)):
            # if 'time' in (str(df[col].dtype)) or 'date' in (str(df[col].dtype)):
            try:
              cell_val = str(try_extract_date(cell_val))

            except:
              print(f'\tno date could be extraceted! attempted val = {cell_val}')

          row_as_dict[col] = cell_val
          c += 1

        dat_dict[r] = row_as_dict
        r += 1

      workbook_dict[s] = dat_dict
      s += 1





  return workbook_dict, sheet_names


# .........................................


def write_to_excel_pyxl_multisheet(workbook_dict, sheet_names_dict, output_file_name):

  #  data format
  """ 1    <--  sheet number

	        {1: {'Client': 'MD, FACS',  'Total Calls': 20}, 2: {'Client': 'Physicians', 'Total Calls': 45} }  <-- rows

  """
  wb1 = openpyxl.workbook.Workbook()

  for sheet_num, data_rows in workbook_dict.items():

    sheet_name = sheet_names_dict[sheet_num]

    ws = wb1.create_sheet(sheet_name)
    ws.title = sheet_name


    headers = [h for h in data_rows[1].keys()]

    ws.append(headers)

    for rnum, row_dict in data_rows.items():
      row_as_list = list(row_dict.values())
      ws.append(row_as_list)

  print('sheet empty --> ', wb1.get_sheet_names()[0])

  if len(wb1.get_sheet_names()) > 1:

    std = wb1.get_sheet_by_name(wb1.get_sheet_names()[0])
    wb1.remove_sheet(std)
    wb1.save(filename=output_file_name)
    print(f'saved to {output_file_name}')


# ...............................................



def excel_file_read_column(filename, sheet_number, col_name, excel_offset=1, headers_row_index=0):
  """ col_name is the column you want its values to be read from the Excel file. Returns those values as a list """
  loc = (filename)

  workbook = xlrd.open_workbook(loc)

  excel_sheet = workbook.sheet_by_index(sheet_number) # read sheet
  #val = excel_sheet.cell_value(0, 0)


  colnames = [] # Reading actual names of excel columns so that we use strings to locate them
  # ASSUMING THAT the first row (index = 0) of Excel sheet is the HEADERS

  for c in range(excel_sheet.ncols):
    cell = excel_sheet.cell_value( headers_row_index, c)
    colnames.append(cell)

  # use colnames produced above to search for the column index using its string value
  i = 0
  possible_col_indexes = []  # in case that column is duplicated, we want the other indexes for it to be recorded
  for c in colnames:
    if col_name.strip() == c.strip():
      possible_col_indexes.append(i) # In excel, it should be the i + 1  column
    i += 1

  col_cell_values = []
  # read values of that column (columns)
  if len(possible_col_indexes) == 1:
    col_index = possible_col_indexes[0]

    # iterate rows
    for row in range(2, excel_sheet.nrows): # start at 1 instead of 0 becasue first row is the column headers and
      # not data
      cell_val = excel_sheet.cell_value(row, col_index)

      # try:
      #       #   a1_as_datetime = datetime.datetime(*xlrd.xldate_as_tuple(cell_val, workbook.datemode))
      #       #   print(a1_as_datetime)
      #       # except Exception as e:
      #       #   print(e)

      col_cell_values.append(cell_val)

    print(possible_col_indexes)
    return col_cell_values, workbook # workbook (needed for date parsing)

  else:
    print(
      'check the len of possible_col_indexes. This error ususally means that column name was not ',
      'found and hence the index of it was not found as well ',
      '  possible_col_indexes ->  ',
          possible_col_indexes)
    return col_cell_values, workbook


# .....................................


def excel_file_loop_rows(filename, sheet_number, start_row_num=1, excel_offset=1, headers_row_index=0):
  """ col_name is the column you want its values to be read from the Excel file. Returns those values as a list """
  loc = (filename)

  workbook = xlrd.open_workbook(loc)

  excel_sheet = workbook.sheet_by_index(sheet_number) # read sheet
  #val = excel_sheet.cell_value(0, 0)

  final_rows = []

  colnames = [] # Reading actual names of excel columns so that we use strings to locate them
  # ASSUMING THAT the first row (index = 0) of Excel sheet is the HEADERS


  for c in range(excel_sheet.ncols):
    cell = excel_sheet.cell_value(headers_row_index, c)
    colnames.append(cell)


  # iterate rows
  for row in range(start_row_num, excel_sheet.nrows):  # start at 1 instead of 0 becasue first row is the column headers and
    # not data

    new_row = {}

    for c in range(excel_sheet.ncols):

      column_name = colnames[c]

      cell_val = excel_sheet.cell_value(row, c)
      if column_name == 'patient_number':
        cell_val = int(float(cell_val))

      new_row[column_name] = cell_val

    print(new_row)
    final_rows.append(new_row)

  write_to_csv('rows_output.csv', field_names=colnames, rows_dictionary=final_rows)

  return final_rows, workbook

# ......................................
def read_csv(filename):
  with open(filename, mode='r') as csv_file:
    csv_reader = csv.reader(csv_file, delimiter=',')
    line_count = 0
    for row in csv_reader:

      print(row)

      line_count += 1

# ..........................

def read_dict_csv(filename, encoding='utf-8'):
  return_values = []
  with open(filename, mode='r', encoding=encoding) as csv_file:
    reader = csv.DictReader(csv_file)

    for row in reader:
        return_values.append(row)

  return return_values


#................................

def write_to_csv(filename, field_names, rows_dictionary):
  with open(filename, mode='w', newline='') as csv_file:
    fieldnames = field_names

    writer = csv.DictWriter(csv_file, fieldnames=fieldnames)

    writer.writeheader()

    for row in rows_dictionary:
      writer.writerow(row)

# .......................................

# def write_to_csv_(filename, field_names, rows_dictionary):
#   with open(filename, mode='w', newline='') as csv_file:
#     fieldnames = field_names
#
#     writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
#
#     writer.writeheader()
#
#     for row in rows_dictionary:
#       writer.writerow(row)

# .................................

def read_text(filename):
  f = open(filename, "r")
  for line in f.readlines():
    print(line)

  return


# .................................................

def get_possible_column_values_in_all_files(col_name):
  file_list = [
     "/excel_files/CopyOf_mhhs_leg_ulcer_and_rightsided_heart_Oct2015-Dec2020.xlsx",
  ]
  sheet_num = 0
  result_dic = {}
  # Populate result_dic. This is in case you need a return value from this method rather than printing
  for file_name in file_list:
   try:
      colvalues, wb = excel_file_read_column(file_name, sheet_num, col_name)
      print(file_name, '\t\t', col_name)
      counter = Counter(colvalues)

      print('\t\t', counter.most_common(), '\n')
      result_dic[file_name] = counter.most_common()
   except Exception as e:
      print('File %s could not be processed!!!'%(file_name))
      print(e.message)
  return result_dic

# .....................................................

# Count frequency of data types in a single column
def get_possible_column_datatypes_in_all_files(col_name):
  file_list = [
    "/excel_files/CopyOf_mhhs_leg_ulcer_and_rightsided_heart_Oct2015-Dec2020.xlsx",

  ]
  sheet_num = 0
  result_dic = {}
  # Populate result_dic. This is in case you need a return value from this method rather than printing
  for file_name in file_list:
    try:
      colvalues, wb = excel_file_read_column(file_name, sheet_num, col_name)
      col_datatypes = [type(x) for x in colvalues]
      print(file_name, '\t\t', col_name)
      counter = Counter(col_datatypes)

      print('\t\t', counter, '\n')
      result_dic[file_name] = counter
    except:
      print('File %s could not be processed!!!'%(file_name))

# ................................

# print values of certain class
def print_values_of_specific_class(list_of_values, __class__):
  return_values = []
  for v in list_of_values:
    if isinstance(v, __class__):
      print(v, '->  ', __class__, '')
      return_values.append(v)

  return return_values

#.............................

def print_dict(dic):
  for v in dic:
    print(v, ' ---> ', dic[v])


#................................



def is_empty_val(cell_val, workbook=None):
  if isinstance(cell_val, str):
    m = re.search(r'^\s+$', cell_val)
    if cell_val == '' or m:
      return True
  return False


# ................................


