import os
import csv
import re
import json
import django
import pandas as pd
import openpyxl
import matplotlib
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd

import matplotlib_funcs as mat_funcs

matplotlib.use('Agg')

import io
import pdb

import dateutil.parser
from datetime import datetime, time, timedelta
import string

from django.conf import settings
from project1.models import TabularDataSets, MyCsvRow


nt = '\n\t'
nl = '\n'

def prepare_data(file_name, dataset_name):
    
    print('\n\n\n')
    
    # file_path = os.path.join(settings.BASE_DIR, file_name)

    # do a file extension check
    extension = os.path.splitext(str(file_name))[1]

    print(f'\n >>> extension: {extension} \n')
    
    if extension.strip() in ['.xls', '.xlsx']:
        headers, rows = read_excel(file_name)
       
        return headers, rows

    
    # reader = csv.reader(file_name)
    contents = file_name.read().decode('UTF-8').splitlines()


    reader = csv.DictReader(contents)
    rows = []

    i = 0
    
    csv_row_headers = []
    for csv_row in reader:

        if  len(csv_row_headers) != len(csv_row.keys()):
            for key in csv_row.keys():
                csv_row_headers.append(key)
        
        rows.append(csv_row)
        i +=1

    new_rows = []
    for csv_row in rows:
        new_row = {}
        for header in csv_row_headers:
            if header != None and header != '':
                new_row[header] = csv_row[header]

        new_rows.append(new_row)
    
    
    return csv_row_headers, rows

# ..............................
            
# ......................................


def read_excel(file_name):
    wb = openpyxl.load_workbook(file_name)
    print(f'\n\t wb = {wb} \n')
    sheet_obj = wb.active
    cell_obj = sheet_obj.cell(row = 2, column = 1)
        
    
    print(f'{nt} cell_obj -> {(cell_obj.value)} {nl}')

    print(f'{nt} cell_obj -> {type(cell_obj.value)} {nl}')

    rows_list = []
    headers = []

    for col in sheet_obj.iter_cols():
        
        headers.append(col[0].value)

    row_num = 0
    for row in sheet_obj.iter_rows():
        if row_num > 0:
            single_row = {}
            c = 0
            for col in headers:
                cell_value = row[c].value
                single_row[col] = cell_value
                c += 1
            
            rows_list.append(single_row)
        row_num += 1
    
    # timedelta_obj = rows_list[2][0]
    # f = convert_timedelta(timedelta_obj, desired_unit='hours')

    new_row_list = []
    for old_row in rows_list:
        new_row = {}
        for col in headers:
            if col != None and col != '':
                new_row[col] = old_row[col]
        new_row_list.append(new_row)

    headers = [h for h in headers if h != None and h !='']

    return headers, new_row_list


#.......................

# .........................................

def detect_datatypes(csv_row_headers, rows):
    #  int - string - float
    dtypes_final_values={}

    dtypes= {}
    for col in csv_row_headers:
        dtypes[col] = []
        for row in rows:

            s = str(row[col]).strip()
            val_type = "string"
            
            x = re.match(r'^[-+]?(\.[0-9]+|[0-9]+\.[0-9]+)$', s)
            
            # Check if float
            if re.match(r'^[-+]?(\.[0-9]+|[0-9]+\.[0-9]+)$', s): 
                val_type = "float"

            # Check if int
            elif re.match(r'^[-+]?[0-9]+$', s):
                val_type = "int"
            elif s == '':
                val_type = "null"
            elif isinstance(row[col], timedelta):
                val_type = 'datetime.timedelta'
            else:
                val_type = "string"

            date_pat1= r'(\d{1,2}(/|-|\\)\d{1,2}(/|-|\\)\d{1,2})|(\d{1,4}(/|-|\\)\d{1,2}(/|-|\\)\d{1,2})|(\d{1,2}(/|-|\\)\d{1,2}(/|-|\\)\d{1,4})'
            #11-Mar-99
            date_word_pat1 = r'(\d+(/|-|\\)\w{3,8})(/|-|\\)\d{2,4}'
            #12-Mar
            date_word_pat2 = r'(\d+(/|-|\\)\w{3,8})'  
            #March-99
            date_word_pat3 = r'(\w{3,8}(/|-|\\)\d+)'
            #15-Mar-1999
            date_word_pat4 = r'\d{1,2}(/|-|\\)(\w{3,8}(/|-|\\)\d+)'
            

            # 23:00 or 1:59
            time_pat1 = r'^((2[0-2])|([0-1]\d)|\d)(\:)[0-5]\d$'
            # 55:35
            time_pat2 = r'^([0-5]\d|\d)\:([0-5]\d)$'
            #55:00:00
            time_pat3 = r'^([0-5]\d|\d)\:([0-5]\d)\:([0-5]\d)$'
            
           

            date_pats = [date_pat1, date_word_pat1, date_word_pat2, date_word_pat3,
                         date_word_pat4]
            time_pats = [time_pat1, time_pat2, time_pat3]
            

            if val_type == "string":
                for date_pat in date_pats:
                    x = re.match(date_pat, s)
                    if x:
                        print(s)
                        try:
                            yourdate = dateutil.parser.parse(s)
                            val_type = 'date'
                            break
                        except:
                            val_type = 'string'
                            print('no date detected by dateutil')

                        # if date_pat == date_word_pat2:
                        #     yourdate.strftime("%b")
                        # elif date_pat == date_word_pat3:
                        #     yourdate.strftime("%B")

                for time_pat in time_pats:
                                        
                    x = re.match(time_pat, s)
                    if x:
                        print(s)
                        if time_pat == time_pat3 or time_pat2:
                                val_type = 'time'

                        try:
                            yourdate = dateutil.parser.parse(s)

                            val_type = 'time'
                            break

                        except:
                            if time_pat == time_pat3 or time_pat2:
                                val_type = 'time'
                            else:
                                val_type = 'string'
                                print('no date detected by dateutil')


                        # if time_pat == time_pat2:
                        #     yourdate.strftime("%H")
                        # elif time_pat == time_pat3:
                        #     yourdate.strftime("%M")
                        

            dtypes[col].append(val_type)



    for col in dtypes.keys():
        print(col, '\n\t', dtypes[col])

        dtypes_final_values[col] = 'string'
        types = dtypes[col]
        

        string_proportions= types.count('string') / len(types)
        float_proportions = types.count('float') / len(types)
        int_proportions= types.count('int') / len(types)
        date_proportions = types.count('date') / len(types)
        time_proportions = types.count('time') / len(types)
        timedelta_proportions = types.count('datetime.timedelta') / len(types)

        
        # float case
        if string_proportions < (float_proportions + int_proportions) and (
            (float_proportions + int_proportions > date_proportions) and 
            (float_proportions + int_proportions > time_proportions) and(
            float_proportions > 0
            )
        ):
            dtypes_final_values[col] = 'float'
            # ---->>>>  nullify the rest and keep numbers only -----
            
        
        # int case
        elif (int_proportions - string_proportions) > 0.01 and (
            int_proportions > date_proportions) and (
            int_proportions > time_proportions 
            ) and (float_proportions <= 0 ):
            dtypes_final_values[col] = 'int'
            # ---->>>>  nullify the rest and keep numbers only -----
            


        # date case
        elif date_proportions > string_proportions and (
            date_proportions > int_proportions
            ) and (date_proportions > float_proportions):

                dtypes_final_values[col] = 'date'
                
        
       #time case
        elif time_proportions > string_proportions and (
            (time_proportions > int_proportions) and (time_proportions > float_proportions)
            and (time_proportions >  date_proportions) ):
                
                dtypes_final_values[col] = 'time'

        # delta-time case
        elif ((timedelta_proportions > time_proportions) and (
                timedelta_proportions > float_proportions) and (
                timedelta_proportions > int_proportions) and (
                timedelta_proportions > string_proportions) and (
                timedelta_proportions > date_proportions
                )):
                dtypes_final_values[col] = 'datetime.timedelta'
        
        else:
            if string_proportions > 0:
                dtypes_final_values[col] = 'string'

       
    print(nl, nl)
    for k, v in dtypes_final_values.items():
        print(k, '\n\t', v, ' ', type(v))

    return dtypes_final_values

#........................

def convert_timedelta(cell_value, desired_unit='hours'):

    final_value = cell_value

    if isinstance(cell_value, timedelta):

        timedelta_obj = cell_value
        days = timedelta_obj.days
        seconds = timedelta_obj.seconds

        hours = seconds//3600
        minutes = (seconds//60)

        
        if desired_unit == 'days':
            final_value = days + (hours/24)

        elif desired_unit == 'hours':
            final_value = days *24 + (hours)
        
        elif desired_unit == 'minutes':
            final_value = days *1440 + (minutes)
    else:
        final_value = None
    
    return final_value

# ................................


def post_process_dtypes(dtypes_values, headers, rows, timedelta_mode='auto'):
    modded_rows = []
    new_dtypes_values = dtypes_values.copy()

    for row in rows:
        modded_rows.append({})

    date_pat1= r'((\d{1,2}(/|-|\\)\d{1,2}(/|-|\\)\d{4}))|((\d{1,4}(/|-|\\)\d{1,2}(/|-|\\)\d{1,2}))|((\d{1,2}(/|-|\\)\d{1,2}(/|-|\\)\d{1,2}))'
    #11-Mar-99
    date_word_pat1 = r'(\d+(/|-|\\)\w{3,8})(/|-|\\)\d{2,4}'
    #12-Mar
    date_word_pat2 = r'(\d+(/|-|\\)\w{3,8})'  
    #March-99
    date_word_pat3 = r'(\w{3,8}(/|-|\\)\d+)'
    #15-Mar-1999
    date_word_pat4 = r'\d{1,2}(/|-|\\)(\w{3,8}(/|-|\\)\d+)'


    # 23:00 or 1:59
    time_pat1 = r'((2[0-2])|([0-1]\d)|\d)(\:)[0-5]\d'
    # 55:35
    time_pat2 = r'^([0-5]\d|\d)\:([0-5]\d)'
    # 55:00:00
    time_pat3 = r'^([0-5]\d|\d)\:([0-5]\d)\:([0-5]\d)'

    col_indx = 0
    for col, dtype in dtypes_values.items():

        # if col=='event_date':
        #     pdb.set_trace()

        row_index = 0
        for row in rows:
            old_val = row[col]
            new_val = None
            
            if dtype == 'float':
                try:
                    new_val = float(old_val)
                except:
                    try:
                        new_val = float(old_val.strip(string.ascii_letters))
                    except:
                        new_val = None
                    
            
            elif dtype =='int':
                try:
                    new_val = int(float(old_val))
                except:
                    try:
                        new_val = int(old_val.strip(string.ascii_letters))
                    except:
                        new_val = None

            elif dtype =='date':
                if 'datetime.datetime' in str(type(old_val)):
                    new_val = old_val
                else:    
                    x2 = re.match(date_word_pat2, str(old_val))
                    x3 = re.match(date_word_pat3, str(old_val))
                    try:
                        yourdate = dateutil.parser.parse(str(old_val))
                        #12-Mar
                        if x2:
                            new_val = time(1,  yourdate.month, yourdate.day)
                        
                        #March-99
                        elif x3:
                            new_val = datetime(yourdate.year, yourdate.month, 1)
                        
                        else:
                            new_val = yourdate

                    except:
                        new_val = None



            elif dtype == "time":
                
                # 23:00 or 1:59
                x2 = re.match(time_pat1, str(old_val))
                # 55:35
                x3 = re.match(time_pat2, str(old_val))
                x4 = re.match(time_pat3, str(old_val))

                try:
                    yourdate = dateutil.parser.parse(old_val)
                    # 23:00 or 1:59
                    if x2:                        
                        new_val = time(yourdate.hour, yourdate.minute, yourdate.second)
                    
                   
                    print(' -> new_val -> ', new_val)

                except:
                     # 55:35
                    if x3:                        
                                                
                        mins = float(x3.group(1))
                        secs = float(x3.group(2)) / 60

                        # return a float value representing minutes as float
                        new_val = mins + secs
                        new_dtypes_values[col] = 'float'

                    # 55:00:00
                    elif x4:
                        
                        gs = [x4.group(1), x4.group(2), x4.group(3)]
                        hours = float(gs[0])
                        mins = float(gs[1]) / 60
                        secs = float(gs[2]) / 3600

                        # return a float value representing hours as float
                        new_val = hours + mins + secs
                        new_dtypes_values[col] = 'float'

                    else:
                        new_val = None

            elif 'datetime.timedelta' in dtype: # because you can have datetime.timedelta hours as dtype (manual change)
                
                if timedelta_mode == 'auto':
                    new_val = convert_timedelta(old_val, desired_unit='hours')
                    new_dtypes_values[col] = 'float'
                elif timedelta_mode == 'manual':
                    desired_unit = dtype.split(' ')[1]
                    new_val = convert_timedelta(old_val, desired_unit)
                    new_dtypes_values[col] = 'float'
                    

            elif dtype == "string":
                new_val = str(old_val)
                # new_val = f"{new_val}"
                           
            modded_rows[row_index][col] = new_val
            row_index += 1

        col_indx += 1
    

   
    
    return modded_rows, new_dtypes_values
          


# ........................

def manual_change_data_type(dtypes_values, target_change_cols, headers, rows):
    # target_change_cols = {colname : 'string'}
    new_dtypes_values = dtypes_values.copy()
    print('\nnew_dtypes_values: \n', new_dtypes_values)
    
    timedelta_mode = 'auto'
    for col, new_dtype in target_change_cols.items():
        # if dtypes_values[col] == 'datetime.timedelta':
        #  new_dtypes for timedelta:  datetime.timedelta days,
        #  datetime.timedelta hours, 
        # datetime.timedelta minutes, datetime.timedelta string (as is)
        if 'datetime.timedelta' in new_dtype:
            timedelta_mode = 'manual'

        new_dtypes_values[col] = new_dtype
        

    print('\nnew_dtypes_values: \n', new_dtypes_values, '\n')
    
    modded_rows, ndtypes = post_process_dtypes(new_dtypes_values, headers, rows, 
                                      timedelta_mode)
    
    return modded_rows, ndtypes

# ................................

def stringfy_data(dtypes_values, headers, rows, timedelta_mode='auto', timedelta_str=None):
    modded_rows = []
    new_dtypes_values = dtypes_values.copy()

    for row in rows:
        new_row = {}
        for col in headers:
            new_row[col] = str(row[col])
            
            if 'datetime.timedelta' in dtypes_values[col]:
                new_row[col] = str(row[col]) + ' hours'
            if timedelta_str:
                new_row[col] = str(row[col]) + ' ' + timedelta_str
        modded_rows.append(new_row)

    return modded_rows

# .....................................

def stringfy_data2(current_dtypes_values, headers, rows):
    modded_rows = []
    new_dtypes_values = current_dtypes_values.copy()

    for row in rows:
        new_row = {}
        for col in headers:
            new_row[col] = str(row[col])
            
            
        modded_rows.append(new_row)

    return modded_rows


# .............................


def convert_to_readable_dtype_value(dtypes_values):
    #  a more readable wording of the data types
    mapper ={'float': 'Decimal number',
             'int': 'Whole number',
             'string': 'Text',
             'date': 'Date', 
             'time': 'Time',
             'datetime.timedelta': 'Time period',}
    readable_dtypes_values = {}
    for col, dtype in dtypes_values.items():
        readable_dtypes_values[col] = mapper[dtype]
    
    return readable_dtypes_values
# ....................................

def reverse_readable_dtype_value(value):
    
    reverse_mapper ={'Decimal number': 'float',
              'Whole number':'int',
              'Text':'string',
              'Date':'date', 
              'Time':'time',
              'Time period':'datetime.timedelta hours'}

        
    return reverse_mapper[value]



# ...........................
def serialize_data(rows):
    # This is a dtype_values agnostic method (serializing is based on individual cells)
    #  seralize both original rows and modded ones
    serialized_data = []

    
    for row in rows:
        new_row = {}
        for each_col, each_val in row.items():
            this_dtype = str(type(each_val))
            new_val = each_val

            if 'datetime.timedelta' in this_dtype:
                new_val = {'type': 'datetime.timedelta', 'days': each_val.days, 
                            'seconds': each_val.seconds}

            elif 'datetime.datetime' in this_dtype:
                new_val = {'type': 'datetime.datetime', 
                            'year': each_val.year,
                            'month': each_val.month, 'day': each_val.day, 
                            }
                try:
                    new_val['hour'] = each_col.hour
                except Exception as e:
                    print(e)
                try:
                    new_val['minute'] = each_col.minute
                except Exception as e:
                    print(e)
                try:
                    new_val['second'] = each_col.second
                except Exception as e:
                    print(e)
            

            new_row[each_col] = new_val
        serialized_data.append(new_row)

    serialized_data_json = json.dumps(serialized_data);
    
    return serialized_data_json

# ...................................


def deserialize_data(json_row_data):
    des_nrows=json.loads(json_row_data)
    new_deserialized_results = []

    for des_row in des_nrows:
        
        new_row = {}
        
        for col, val in des_row.items():
            
            t = str(type(val))
                        
            if 'dict' in t:
                if 'type' in val:
                    if 'datetime.timedelta' in val['type']:
                        nval = timedelta(days=val['days'], seconds=val['seconds'])
                        new_row[col] = nval
                    elif 'datetime.datetime' in val['type']:
                        nval = datetime(year=val['year'], month=val['month'],
                                        day=val['day'])
                        new_row[col] = nval
            else:
                new_row[col] = val

        new_deserialized_results.append(new_row) 

    return new_deserialized_results



# ..............................................

def determine_allowed_xy(chart_type, modded_rows, headers, current_dtype_values):
    allowed_x_names = []
    allowed_y_names = []

    allowed_x_types = []
    allowed_y_types = []
    
    # 'datetime.timedelta' 
    if chart_type == 'BAR-CHART':
        # x can be categorical ('strings')
        for col, dtype in current_dtype_values.items():
            if 'string' in dtype or 'int' in dtype or 'date' in dtype:
                allowed_x_names.append(col)

        allowed_x_types =['Text', 'Number', 'Date']
        allowed_y_types = ['Number']               

            
    if chart_type in ['LINE-CHART', 'MULTI-LINE-CHART']:
        for col, dtype in current_dtype_values.items():
            if 'float' in dtype or 'int' in dtype or 'date' in dtype:
                allowed_x_names.append(col)
                

        allowed_x_types =['Number', 'Date']
        allowed_y_types = ['Number']

    
    if chart_type == 'PIE-CHART':
        # x can be categorical ('strings')
        for col, dtype in current_dtype_values.items():
            if 'string' in dtype:
                allowed_x_names.append(col)
                
        allowed_x_types =['Text']
        allowed_y_types = ['Number']


    if chart_type == 'HISTOGRAM':
        # x can be categorical ('strings')
        for col, dtype in current_dtype_values.items():
            if 'int' in dtype or 'float' in dtype:
                allowed_x_names.append(col)
        
        allowed_x_types =[ 'Number']
        allowed_y_types = ['']

    if chart_type == "SCATTERPLOT":
        for col, dtype in current_dtype_values.items():
            if 'float' in dtype or 'int' in dtype:
                allowed_x_names.append(col)

    for col, dtype in current_dtype_values.items():
        if dtype in ['int', 'float']:
            allowed_y_names.append(col)
        
    
        

    return allowed_x_names, allowed_y_names, allowed_x_types, allowed_y_types

# ................................

def sort_dates(column_to_sort, modded_rows, headers, current_dtype_values):
    sorted_modded_rows = []
    
    sorted_modded_rows = sorted(modded_rows, key = lambda data: data[column_to_sort])


    return sorted_modded_rows

# ...............

def stringify_dates(modded_rows, current_dtype_values):
    new_modded_rows = []
    for mr in modded_rows:
        new_mr = {}

        for col, dtype in current_dtype_values.items():
            if dtype == 'date':
                new_val = mr[col]
                if new_val:
                    new_val = mr[col].strftime("%m-%d-%Y")
                    new_val = f"{new_val}"
                new_mr[col] = new_val
            else:
                new_mr[col] = mr[col]

        new_modded_rows.append(new_mr)        


    return new_modded_rows

# ............................
# =====================================================

# ----> VIZUALIZATIONS <-----


def multiline_chart(rows, dtypes_values, x_name, y_names):
    mat_funcs.multiline_chart(rows, dtypes_values, x_name, y_names)
# .........
def single_line_chart(rows, x_name, y_values, dtypes_values=None):

    mat_funcs.single_line_chart(rows, x_name, y_values, dtypes_values)

# .........
def single_scatter_plot(rows, x_name, y_values, dtypes_values=None, trend_line=True):
    mat_funcs.single_scatter_plot(rows, x_name, y_values, dtypes_values, trend_line)

# .........

def histogram(rows, x_name, y_values, dtypes_values=None):
    mat_funcs.histogram(rows, x_name, y_values, dtypes_values)

# ..........

def pie_chart(rows, x_name, y_values, dtypes_values=None):
    mat_funcs.pie_chart(rows, x_name, y_values, dtypes_values)

# .........

def bar_chart(rows, x_name, y_values, dtypes_values=None):
    mat_funcs.bar_chart(rows, x_name, y_values, dtypes_values)





# >>>>>>>>> TEST THE PIPELINE <<<<<<<<<<<<<<

# 1 READ EXCEL (test excel parse will be separate)
# 2 put timedelta
# 3 print(dtypes) and rows
# 4 TEST manual change then print modded rows


# ............................
# ...........................
# Deprecate?
def post_process_excel_rows(rows_list, conversion, desired_unit='hours'):
    new_rows_list = []
    for row in rows_list:
        print(row)
        new_row = {}
        for colname, cell in row.items():
            new_cell = cell
            if conversion == 'timedelta_conversion':
                new_cell = convert_timedelta(cell, desired_unit)
            
            new_row[colname] = new_cell

        new_rows_list.append(new_row)
    
    return new_rows_list

# ................................
# To be Deprecated?
def unpack_csvrow_values(row_values, separator='<,>'):

    '2011<,>45'
    
    result_row = []
    separator = r'(?!"){}(?!")'.format(separator)

    escaped = '"{}"'.format(separator)



    a = re.split(separator, row_values)

    b = [re.sub(r'"<,>"', '<,>', x) for x in a]  # remove escapes

    result_row = b

    return result_row


#................................

def convert_rows_to_json(column_names: list , list_of_rows: list):
    py_major_dict = {}

    n = 0
    for row in list_of_rows:
        d = {}
        for i in range(0, len(row)):
            d [column_names[i]] = row[i]

        py_major_dict[n] = d

        n += 1
    
    return json.dumps(py_major_dict)



